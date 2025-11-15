import os
import re
import time
import logging
import pandas as pd
from urllib.parse import quote_plus
from sqlalchemy import create_engine, text
from sqlalchemy.exc import SQLAlchemyError

# ---------- Configuración ----------
DB_USER = os.getenv('DB_USER', 'root')
DB_PASS = os.getenv('DB_PASS', 'dominid')
DB_HOST = os.getenv('DB_HOST', 'localhost')
DB_PORT = os.getenv('DB_PORT', '3306')
DB_NAME = os.getenv('DB_NAME', 'andinoVuelo')

DATA_DIR = "data"
EXCEL_FILE = os.path.join(DATA_DIR, "plantillas_andino.xlsx")

# Tablas del esquema (deben existir CSV con estos nombres)
CSV_FILES = {
    "aerolinea":     os.path.join(DATA_DIR, "aerolinea.csv"),
    "aeronave":      os.path.join(DATA_DIR, "aeronave.csv"),
    "aeropuerto":    os.path.join(DATA_DIR, "aeropuerto.csv"),
    "terminal":      os.path.join(DATA_DIR, "terminal.csv"),
    "puerta":        os.path.join(DATA_DIR, "puerta.csv"),
    "vuelo":         os.path.join(DATA_DIR, "vuelo.csv"),
    "pasajero":      os.path.join(DATA_DIR, "pasajero.csv"),
    "ticket_aereo":  os.path.join(DATA_DIR, "ticket_aereo.csv"),
    "pase_abordar":  os.path.join(DATA_DIR, "pase_abordar.csv"),
    "equipaje":      os.path.join(DATA_DIR, "equipaje.csv"),
    "embarque":      os.path.join(DATA_DIR, "embarque.csv"),
    "log_cambios":   os.path.join(DATA_DIR, "log_cambios.csv"),
}

OUT_DIR = "output"
OUT_FILE = os.path.join(OUT_DIR, f"diccionario_{DB_NAME}_completo.xlsx")
OUT_TABLES_DIR = os.path.join(OUT_DIR, "tables")
SAVE_CSV_PER_TABLE = True

# Logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# Engine SQLAlchemy
ENC_PASS = quote_plus(DB_PASS) if DB_PASS else ''
engine = create_engine(
    f"mysql+pymysql://{DB_USER}:{ENC_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}?charset=utf8mb4",
    pool_pre_ping=True,
    pool_recycle=3600
)

# Cache de columnas por tabla
TABLE_COLS_CACHE = {}

def get_table_columns(conn, table):
    if table not in TABLE_COLS_CACHE:
        rows = conn.execute(text("""
            SELECT COLUMN_NAME
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = :schema AND TABLE_NAME = :table
        """), {"schema": DB_NAME, "table": table}).scalars().all()
        TABLE_COLS_CACHE[table] = set(rows)
    return TABLE_COLS_CACHE[table]

# ---------- Helpers ----------
def read_sheet_or_csv(name, sheet_name=None):
    csv_path = CSV_FILES.get(name)
    if csv_path and os.path.exists(csv_path):
        try:
            df = pd.read_csv(csv_path, dtype=str, keep_default_na=False, na_values=[''])
            log.info("✓ Leído CSV %s (%d filas)", os.path.basename(csv_path), len(df))
            return df
        except Exception as e:
            log.warning("Error leyendo CSV %s: %s", csv_path, e)
    
    if os.path.exists(EXCEL_FILE):
        try:
            sheet = name if sheet_name is None else sheet_name
            df = pd.read_excel(EXCEL_FILE, sheet_name=sheet, engine='openpyxl', dtype=str)
            log.info("✓ Leído hoja '%s' de Excel (%d filas)", sheet, len(df))
            return df
        except Exception as e:
            log.warning("No se pudo leer hoja '%s': %s", name, e)
    
    return pd.DataFrame()

def normalize_df(df):
    if df is None or df.empty:
        return pd.DataFrame()
    return df.replace({'': None}).where(pd.notnull(df), None)

def _to_int_safe(val):
    if val is None:
        return None
    if isinstance(val, int):
        return val
    try:
        s = str(val).strip()
        if s == '':
            return None
        return int(float(s))
    except Exception:
        return None

def _to_bool_safe(val):
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ('1','true','t','yes','y','si','sí'):
        return True
    if s in ('0','false','f','no','n'):
        return False
    return None

def table_exists(conn, table_name):
    q = text("""
      SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES
      WHERE TABLE_SCHEMA = :schema AND TABLE_NAME = :table
    """)
    return bool(conn.execute(q, {"schema": DB_NAME, "table": table_name}).scalar())

def insert_row_if_missing(conn, table, row, pk='id'):
    valid_cols = get_table_columns(conn, table)
    row = {k: (None if pd.isna(v) else v) for k, v in row.items() if k in valid_cols}
    if pk in row:
        row[pk] = _to_int_safe(row[pk])

    cols_nonnull = {k: v for k, v in row.items() if v is not None}
    if not cols_nonnull:
        return False

    cols = list(cols_nonnull.keys())
    vals = {c: cols_nonnull[c] for c in cols}
    cols_sql = ", ".join(f"`{c}`" for c in cols)
    placeholders = ", ".join([f":{c}" for c in cols])

    if pk in cols_nonnull:
        exists = conn.execute(
            text(f"SELECT 1 FROM {table} WHERE {pk} = :id LIMIT 1"),
            {"id": cols_nonnull[pk]}
        ).fetchone()
        if exists:
            return False

    sql = f"INSERT INTO {table} ({cols_sql}) VALUES ({placeholders})"
    conn.execute(text(sql), vals)
    return True

def load_optional_table(conn, table_name, df, fk_checks):
    if not table_exists(conn, table_name):
        log.warning("⊘ Se omite '%s' (tabla no existe)", table_name)
        return 0, 0, 0, pd.DataFrame()

    for _, ref_table in fk_checks.items():
        if not table_exists(conn, ref_table):
            log.warning("⊘ Se omite '%s' (tabla referenciada '%s' no existe)", table_name, ref_table)
            return 0, 0, 0, pd.DataFrame()

    inserted = 0
    updated = 0
    skipped = 0
    invalid_rows = []
    fk_cache = {}

    def exists_in(table, id_):
        if id_ is None:
            return False
        if table not in fk_cache:
            rows = conn.execute(text(f"SELECT id FROM {table}")).fetchall()
            fk_cache[table] = {r[0] for r in rows}
        return id_ in fk_cache[table]

    def exists_in_dest(id_):
        if id_ is None:
            return False
        res = conn.execute(text(f"SELECT 1 FROM {table_name} WHERE id = :id LIMIT 1"), {"id": id_}).fetchone()
        return bool(res)

    valid_cols_dest = get_table_columns(conn, table_name)

    for _, r in df.iterrows():
        row = {k: v for k, v in r.to_dict().items() if k in valid_cols_dest}

        for fk_col, _ in fk_checks.items():
            if fk_col in row:
                row[fk_col] = _to_int_safe(row[fk_col])
        if 'validado' in row:
            b = _to_bool_safe(row['validado'])
            if b is not None:
                row['validado'] = b

        fk_ok = True
        for fk_col, ref_table in fk_checks.items():
            if fk_col not in row or row.get(fk_col) is None or not exists_in(ref_table, row.get(fk_col)):
                fk_ok = False
                break
        if not fk_ok:
            invalid_rows.append(row)
            continue

        try:
            id_val = _to_int_safe(row.get('id')) if 'id' in row else None
            if id_val is not None and exists_in_dest(id_val):
                cols = [k for k in row.keys() if k != 'id' and row[k] is not None]
                if not cols:
                    skipped += 1
                    continue
                set_sql = ", ".join([f"`{c}` = :{c}" for c in cols])
                params = {c: row[c] for c in cols}
                params['id'] = id_val
                update_sql = f"UPDATE {table_name} SET {set_sql} WHERE id = :id"
                conn.execute(text(update_sql), params)
                updated += 1
            else:
                if insert_row_if_missing(conn, table_name, row, pk='id'):
                    inserted += 1
                else:
                    skipped += 1
        except Exception as e:
            log.warning("Error procesando %s: %s", table_name, e)
            invalid_rows.append(row)

    return inserted, updated, skipped, pd.DataFrame(invalid_rows)

# ---------- ETL principal ----------
def run_etl():
    log.info("=" * 60)
    log.info("INICIANDO ETL - Carga de datos aeroportuarios")
    log.info("=" * 60)
    
    # Leer todas las fuentes
    df_aerolinea    = normalize_df(read_sheet_or_csv("aerolinea"))
    df_aeronave     = normalize_df(read_sheet_or_csv("aeronave"))
    df_aeropuerto   = normalize_df(read_sheet_or_csv("aeropuerto"))
    df_terminal     = normalize_df(read_sheet_or_csv("terminal"))
    df_puerta       = normalize_df(read_sheet_or_csv("puerta"))
    df_vuelo        = normalize_df(read_sheet_or_csv("vuelo"))
    df_pasajero     = normalize_df(read_sheet_or_csv("pasajero"))
    df_ticket       = normalize_df(read_sheet_or_csv("ticket_aereo"))
    df_pase         = normalize_df(read_sheet_or_csv("pase_abordar"))
    df_equipaje     = normalize_df(read_sheet_or_csv("equipaje"))
    df_embarque     = normalize_df(read_sheet_or_csv("embarque"))
    df_logs         = normalize_df(read_sheet_or_csv("log_cambios"))

    # Normalizar ticket_aereo
    if not df_ticket.empty:
        df_ticket = df_ticket.rename(columns={
            'PNR': 'pnr', 'PasajeroID': 'pasajero_id', 'VueloID': 'vuelo_id',
            'Asiento': 'asiento', 'Estado': 'estado_ticket', 'FechaEmision': 'fecha_emision'
        })

    log.info("\n--- FASE 1: Cargando tablas maestras ---")
    
    # 1) Insertar maestros
    try:
        with engine.begin() as conn:
            summary = {
                "aerolinea": 0, "aeronave": 0, "aeropuerto": 0, 
                "terminal": 0, "puerta": 0, "vuelo": 0, "pasajero": 0
            }

            # Aerolíneas
            if not df_aerolinea.empty:
                for _, r in df_aerolinea.iterrows():
                    if insert_row_if_missing(conn, "aerolinea", r.to_dict(), pk='id'):
                        summary["aerolinea"] += 1
                log.info("  ✓ aerolinea: %d registros insertados", summary["aerolinea"])

            # Aeropuertos
            if not df_aeropuerto.empty:
                for _, r in df_aeropuerto.iterrows():
                    if insert_row_if_missing(conn, "aeropuerto", r.to_dict(), pk='id'):
                        summary["aeropuerto"] += 1
                log.info("  ✓ aeropuerto: %d registros insertados", summary["aeropuerto"])

            # Terminales
            if not df_terminal.empty:
                for _, r in df_terminal.iterrows():
                    if insert_row_if_missing(conn, "terminal", r.to_dict(), pk='id'):
                        summary["terminal"] += 1
                log.info("  ✓ terminal: %d registros insertados", summary["terminal"])

            # Puertas
            if not df_puerta.empty:
                for _, r in df_puerta.iterrows():
                    row = r.to_dict()
                    tid = _to_int_safe(row.get('terminal_id'))
                    if tid is not None:
                        exists = conn.execute(text("SELECT 1 FROM terminal WHERE id = :id LIMIT 1"), {"id": tid}).fetchone()
                        if not exists:
                            continue
                        row['terminal_id'] = tid
                    if insert_row_if_missing(conn, "puerta", row, pk='id'):
                        summary["puerta"] += 1
                log.info("  ✓ puerta: %d registros insertados", summary["puerta"])

            # Aeronaves
            if not df_aeronave.empty:
                for _, r in df_aeronave.iterrows():
                    row = r.to_dict()
                    if 'capacidad' in row and 'capacidad_pasajeros' not in row:
                        row['capacidad_pasajeros'] = row.pop('capacidad')
                    if insert_row_if_missing(conn, "aeronave", row, pk='id'):
                        summary["aeronave"] += 1
                log.info("  ✓ aeronave: %d registros insertados", summary["aeronave"])

            # Vuelos
            if not df_vuelo.empty:
                for _, r in df_vuelo.iterrows():
                    row = r.to_dict()
                    for key in ('aerolinea_id','aeronave_id','puerta_id','aeropuerto_origen_id','aeropuerto_destino_id'):
                        if key in row:
                            row[key] = _to_int_safe(row.get(key))
                    
                    # Validar FKs
                    if row.get('aerolinea_id') and not conn.execute(text("SELECT 1 FROM aerolinea WHERE id=:i"), {"i": row['aerolinea_id']}).fetchone():
                        continue
                    if row.get('aeronave_id') and not conn.execute(text("SELECT 1 FROM aeronave WHERE id=:i"), {"i": row['aeronave_id']}).fetchone():
                        continue
                    if row.get('aeropuerto_origen_id') and not conn.execute(text("SELECT 1 FROM aeropuerto WHERE id=:i"), {"i": row['aeropuerto_origen_id']}).fetchone():
                        continue
                    if row.get('aeropuerto_destino_id') and not conn.execute(text("SELECT 1 FROM aeropuerto WHERE id=:i"), {"i": row['aeropuerto_destino_id']}).fetchone():
                        continue
                    
                    if insert_row_if_missing(conn, "vuelo", row, pk='id'):
                        summary["vuelo"] += 1
                log.info("  ✓ vuelo: %d registros insertados", summary["vuelo"])

            # Pasajeros
            if not df_pasajero.empty:
                for _, r in df_pasajero.iterrows():
                    if insert_row_if_missing(conn, "pasajero", r.to_dict(), pk='id'):
                        summary["pasajero"] += 1
                log.info("  ✓ pasajero: %d registros insertados", summary["pasajero"])

    except SQLAlchemyError as e:
        log.exception("❌ Error insertando maestros: %s", e)
        return

    log.info("\n--- FASE 2: Cargando tickets ---")
    
    # 2) Merge ticket_aereo
    if not df_ticket.empty:
        try:
            df_ticket['pasajero_id'] = df_ticket['pasajero_id'].apply(_to_int_safe)
            df_ticket['vuelo_id'] = df_ticket['vuelo_id'].apply(_to_int_safe)
        except Exception:
            log.warning("No se pudo castear IDs en ticket_aereo")

        df_ticket = df_ticket.drop_duplicates(subset=['pnr'])

        with engine.connect() as conn_check:
            if not table_exists(conn_check, 'ticket_aereo'):
                log.warning("⊘ Tabla 'ticket_aereo' no existe")
            else:
                pasajeros_db = {r[0] for r in conn_check.execute(text("SELECT id FROM pasajero")).fetchall()}
                vuelos_db = {r[0] for r in conn_check.execute(text("SELECT id FROM vuelo")).fetchall()}
                df_valid = df_ticket[df_ticket['pasajero_id'].isin(pasajeros_db) & df_ticket['vuelo_id'].isin(vuelos_db)].copy()
                df_invalid = df_ticket[~df_ticket.index.isin(df_valid.index)]
                
                if not df_invalid.empty:
                    out = os.path.join(DATA_DIR, "ticket_aereo_invalidos_fk.csv")
                    df_invalid.to_csv(out, index=False, encoding='utf-8')
                    log.warning("  ⚠ %d tickets inválidos guardados en %s", len(df_invalid), out)

                if not df_valid.empty:
                    ts = int(time.time())
                    staging_table = f"ticket_staging_{os.getpid()}_{ts}"
                    try:
                        # Crear tabla staging con misma collation que ticket_aereo
                        with engine.begin() as conn:
                            # Obtener collation de ticket_aereo.pnr
                            collation_query = text("""
                                SELECT COLLATION_NAME 
                                FROM INFORMATION_SCHEMA.COLUMNS 
                                WHERE TABLE_SCHEMA = :schema 
                                  AND TABLE_NAME = 'ticket_aereo' 
                                  AND COLUMN_NAME = 'pnr'
                                LIMIT 1
                            """)
                            collation_result = conn.execute(collation_query, {"schema": DB_NAME}).fetchone()
                            collation = collation_result[0] if collation_result else 'utf8mb4_unicode_ci'
                            
                            # Crear tabla staging con collation correcta
                            create_staging = text(f"""
                                CREATE TABLE {staging_table} (
                                    pnr VARCHAR(255) COLLATE {collation},
                                    pasajero_id INT,
                                    vuelo_id INT,
                                    asiento VARCHAR(10) COLLATE {collation},
                                    estado_ticket VARCHAR(50) COLLATE {collation},
                                    fecha_emision VARCHAR(50) COLLATE {collation}
                                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE={collation}
                            """)
                            conn.execute(create_staging)
                            log.info("  ✓ Tabla staging creada con collation: %s", collation)
                        
                        # Eliminar columna 'id' si existe (no la necesitamos en staging)
                        df_to_stage = df_valid.copy()
                        if 'id' in df_to_stage.columns:
                            df_to_stage = df_to_stage.drop(columns=['id'])
                        
                        # Insertar datos en staging
                        df_to_stage.to_sql(staging_table, con=engine, if_exists='append', index=False)
                        log.info("  ✓ %d tickets válidos cargados en staging", len(df_to_stage))
                        
                        with engine.begin() as conn:
                            insert_cols = ["pnr", "pasajero_id", "vuelo_id", "asiento", "estado_ticket", "fecha_emision"]
                            
                            # INSERT para tickets nuevos
                            insert_sql = f"""
                            INSERT INTO ticket_aereo ({", ".join(insert_cols)})
                            SELECT {", ".join("ts."+c for c in insert_cols)}
                            FROM {staging_table} ts
                            LEFT JOIN ticket_aereo t ON t.pnr = ts.pnr
                            WHERE t.id IS NULL;
                            """
                            res_ins = conn.execute(text(insert_sql))
                            inserted = getattr(res_ins, "rowcount", 0)
                            
                            # UPDATE para tickets existentes
                            update_sql = f"""
                            UPDATE ticket_aereo t
                            JOIN {staging_table} ts ON t.pnr = ts.pnr
                            SET t.pasajero_id = ts.pasajero_id,
                                t.vuelo_id = ts.vuelo_id,
                                t.asiento = ts.asiento,
                                t.estado_ticket = ts.estado_ticket;
                            """
                            res_upd = conn.execute(text(update_sql))
                            updated = getattr(res_upd, "rowcount", 0)
                            
                            log.info("  ✓ ticket_aereo: %d insertados, %d actualizados", inserted, updated)
                    
                    except Exception as e:
                        log.error("  ❌ Error cargando tickets: %s", e)
                        log.exception(e)
                    
                    finally:
                        # Limpiar staging
                        try:
                            with engine.connect() as cdrop:
                                cdrop.execute(text(f"DROP TABLE IF EXISTS {staging_table}"))
                                cdrop.commit()
                                log.info("  ✓ Tabla staging eliminada")
                        except Exception:
                            pass
    else:
        log.info("  ⊘ No hay datos de ticket_aereo")

    log.info("\n--- FASE 3: Cargando tablas relacionadas ---")
    
    # 3) Importar opcionales
    try:
        with engine.begin() as conn:
            # Pases de abordar
            if not df_pase.empty:
                ins, upd, skp, inv = load_optional_table(conn, "pase_abordar", df_pase, {"ticket_aereo_id": "ticket_aereo"})
                log.info("  ✓ pase_abordar: %d insertados, %d actualizados, %d saltados", ins, upd, skp)
                if not inv.empty:
                    inv.to_csv(os.path.join(DATA_DIR, "pase_abordar_invalidas.csv"), index=False)

            # Equipaje
            if not df_equipaje.empty:
                ins, upd, skp, inv = load_optional_table(conn, "equipaje", df_equipaje, {"ticket_aereo_id": "ticket_aereo", "vuelo_id": "vuelo"})
                log.info("  ✓ equipaje: %d insertados, %d actualizados, %d saltados", ins, upd, skp)
                if not inv.empty:
                    inv.to_csv(os.path.join(DATA_DIR, "equipaje_invalidas.csv"), index=False)

            # Embarque
            if not df_embarque.empty:
                ins, upd, skp, inv = load_optional_table(conn, "embarque", df_embarque, {"vuelo_id": "vuelo", "ticket_aereo_id": "ticket_aereo", "puerta_id": "puerta"})
                log.info("  ✓ embarque: %d insertados, %d actualizados, %d saltados", ins, upd, skp)
                if not inv.empty:
                    inv.to_csv(os.path.join(DATA_DIR, "embarque_invalidas.csv"), index=False)

            # Log de cambios
            if not df_logs.empty:
                if not table_exists(conn, "log_cambios"):
                    log.warning("  ⊘ Tabla 'log_cambios' no existe")
                else:
                    inserted = 0
                    for _, r in df_logs.iterrows():
                        row = r.to_dict()
                        try:
                            if 'id' in row:
                                row['id'] = _to_int_safe(row['id'])
                        except Exception:
                            pass
                        try:
                            if insert_row_if_missing(conn, "log_cambios", row, pk='id'):
                                inserted += 1
                        except Exception:
                            pass
                    log.info("  ✓ log_cambios: %d registros insertados", inserted)
    except Exception as e:
        log.exception("❌ Error importando tablas relacionadas: %s", e)

    log.info("\n" + "=" * 60)
    log.info("ETL COMPLETADO")
    log.info("=" * 60 + "\n")

# ---------- Diccionario Excel ----------
def generar_diccionario():
    log.info("=" * 60)
    log.info("GENERANDO DICCIONARIO DE DATOS")
    log.info("=" * 60)
    
    os.makedirs(OUT_DIR, exist_ok=True)
    if SAVE_CSV_PER_TABLE:
        os.makedirs(OUT_TABLES_DIR, exist_ok=True)
    
    try:
        with engine.connect() as conn:
            # Consultas de metadatos
            SQL_COLUMNS = """
            SELECT
              TABLE_NAME,
              ORDINAL_POSITION,
              COLUMN_NAME,
              COLUMN_TYPE,
              DATA_TYPE,
              CHARACTER_MAXIMUM_LENGTH,
              NUMERIC_PRECISION,
              IS_NULLABLE,
              COLUMN_DEFAULT,
              COLUMN_KEY,
              EXTRA,
              COLUMN_COMMENT
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = :schema
            ORDER BY TABLE_NAME, ORDINAL_POSITION;
            """
            
            SQL_FKS = """
            SELECT
              kcu.TABLE_NAME,
              kcu.COLUMN_NAME,
              kcu.CONSTRAINT_NAME,
              kcu.REFERENCED_TABLE_NAME,
              kcu.REFERENCED_COLUMN_NAME
            FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE kcu
            WHERE kcu.TABLE_SCHEMA = :schema
              AND kcu.REFERENCED_TABLE_NAME IS NOT NULL
            ORDER BY kcu.TABLE_NAME;
            """
            
            SQL_INDEXES = """
            SELECT
              s.TABLE_NAME,
              s.INDEX_NAME,
              s.NON_UNIQUE,
              s.SEQ_IN_INDEX,
              s.COLUMN_NAME,
              s.COLLATION,
              s.SUB_PART
            FROM INFORMATION_SCHEMA.STATISTICS s
            WHERE s.TABLE_SCHEMA = :schema
            ORDER BY s.TABLE_NAME, s.INDEX_NAME, s.SEQ_IN_INDEX;
            """
            
            SQL_TABLES = """
            SELECT
              t.TABLE_NAME AS tabla,
              t.TABLE_ROWS AS filas_aprox,
              ROUND(t.DATA_LENGTH / 1024, 2) AS datos_kb,
              ROUND(t.INDEX_LENGTH / 1024, 2) AS indices_kb,
              ROUND((t.DATA_LENGTH + t.INDEX_LENGTH) / 1024, 2) AS total_kb
            FROM INFORMATION_SCHEMA.TABLES t
            WHERE t.TABLE_SCHEMA = :schema AND t.TABLE_TYPE = 'BASE TABLE'
            ORDER BY t.TABLE_NAME;
            """
            
            log.info("Consultando metadatos de la base de datos...")
            df_columns = pd.read_sql(text(SQL_COLUMNS), conn, params={"schema": DB_NAME})
            df_fks = pd.read_sql(text(SQL_FKS), conn, params={"schema": DB_NAME})
            
            try:
                df_indexes = pd.read_sql(text(SQL_INDEXES), conn, params={"schema": DB_NAME})
            except Exception:
                df_indexes = pd.DataFrame()
            
            df_tables = pd.read_sql(text(SQL_TABLES), conn, params={"schema": DB_NAME})
            
            # Renombrar columnas para mejor legibilidad
            df_columns = df_columns.rename(columns={
                "TABLE_NAME": "tabla",
                "ORDINAL_POSITION": "posicion",
                "COLUMN_NAME": "columna",
                "COLUMN_TYPE": "tipo_completo",
                "DATA_TYPE": "tipo_dato",
                "CHARACTER_MAXIMUM_LENGTH": "long_max",
                "NUMERIC_PRECISION": "precision_num",
                "IS_NULLABLE": "permite_null",
                "COLUMN_DEFAULT": "valor_default",
                "COLUMN_KEY": "tipo_clave",
                "EXTRA": "extra",
                "COLUMN_COMMENT": "comentario"
            })
            
            df_fks = df_fks.rename(columns={
                "TABLE_NAME": "tabla",
                "COLUMN_NAME": "columna",
                "CONSTRAINT_NAME": "nombre_constraint",
                "REFERENCED_TABLE_NAME": "tabla_referenciada",
                "REFERENCED_COLUMN_NAME": "columna_referenciada"
            })
            
            if not df_indexes.empty:
                df_indexes = df_indexes.rename(columns={
                    "TABLE_NAME": "tabla",
                    "INDEX_NAME": "nombre_indice",
                    "NON_UNIQUE": "no_unico",
                    "SEQ_IN_INDEX": "secuencia",
                    "COLUMN_NAME": "columna",
                    "COLLATION": "colacion",
                    "SUB_PART": "sub_parte"
                })
            
            log.info("Generando archivo Excel...")
            
            # Crear Excel con múltiples hojas
            with pd.ExcelWriter(OUT_FILE, engine='openpyxl') as writer:
                # Hoja 1: Resumen de tablas
                df_tables.to_excel(writer, sheet_name="1_Resumen_Tablas", index=False)
                log.info("  ✓ Hoja 'Resumen_Tablas' creada")
                
                # Hoja 2: Todas las columnas
                df_columns.to_excel(writer, sheet_name="2_Todas_Columnas", index=False)
                log.info("  ✓ Hoja 'Todas_Columnas' creada")
                
                # Hoja 3: Foreign Keys
                if not df_fks.empty:
                    df_fks.to_excel(writer, sheet_name="3_Foreign_Keys", index=False)
                    log.info("  ✓ Hoja 'Foreign_Keys' creada")
                
                # Hoja 4: Índices
                if not df_indexes.empty:
                    df_indexes.to_excel(writer, sheet_name="4_Indices", index=False)
                    log.info("  ✓ Hoja 'Indices' creada")
                
                # Hojas 5+: Una por cada tabla
                tables = sorted(df_columns['tabla'].unique().tolist())
                used_sheet_names = {'1_Resumen_Tablas', '2_Todas_Columnas', '3_Foreign_Keys', '4_Indices'}
                
                for idx, tbl in enumerate(tables, start=5):
                    tbl_df = df_columns[df_columns['tabla'] == tbl].copy().sort_values('posicion')
                    
                    # Nombre de hoja limpio (max 31 caracteres)
                    sheet_name = f"{idx}_{tbl}"[:31]
                    sheet_name = re.sub(r'[:\\/?*\[\]]', '_', sheet_name)
                    
                    # Asegurar unicidad
                    base = sheet_name
                    counter = 1
                    while sheet_name in used_sheet_names:
                        suffix = f"_{counter}"
                        allowed = 31 - len(suffix)
                        sheet_name = base[:allowed] + suffix
                        counter += 1
                    used_sheet_names.add(sheet_name)
                    
                    # Eliminar columna 'tabla' (redundante en hoja individual)
                    tbl_df = tbl_df.drop(columns=['tabla'], errors='ignore')
                    
                    # Escribir a Excel
                    tbl_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Guardar CSV individual si está habilitado
                    if SAVE_CSV_PER_TABLE:
                        csv_path = os.path.join(OUT_TABLES_DIR, f"{tbl}.csv")
                        tbl_df.to_csv(csv_path, index=False, encoding='utf-8')
                
                log.info(f"  ✓ {len(tables)} hojas de tablas individuales creadas")
            
            # Aplicar formato y autoajuste de columnas
            log.info("Aplicando formato al archivo Excel...")
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = load_workbook(OUT_FILE)
                
                # Estilo de encabezados
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for ws in wb.worksheets:
                    # Autoajustar anchos de columna
                    dims = {}
                    for row in ws.rows:
                        for cell in row:
                            value = '' if cell.value is None else str(cell.value)
                            dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(value))
                    
                    for col, width in dims.items():
                        adjusted_width = min(max(width + 2, 10), 60)
                        ws.column_dimensions[col].width = adjusted_width
                    
                    # Formatear encabezados (primera fila)
                    if ws.max_row > 0:
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                    
                    # Congelar primera fila
                    ws.freeze_panes = "A2"
                
                wb.save(OUT_FILE)
                log.info("  ✓ Formato aplicado correctamente")
            except ImportError:
                log.warning("  ⚠ openpyxl no disponible para formato avanzado")
            except Exception as e:
                log.warning(f"  ⚠ Error aplicando formato: {e}")
            
            log.info(f"\n✓ Diccionario Excel generado: {OUT_FILE}")
            
            if SAVE_CSV_PER_TABLE:
                log.info(f"✓ CSV individuales guardados en: {OUT_TABLES_DIR}")
            
            # Mostrar estadísticas finales
            log.info("\n--- ESTADÍSTICAS FINALES ---")
            log.info(f"Total de tablas: {len(df_tables)}")
            log.info(f"Total de columnas: {len(df_columns)}")
            log.info(f"Total de FKs: {len(df_fks)}")
            if not df_indexes.empty:
                log.info(f"Total de índices: {df_indexes['nombre_indice'].nunique()}")
            
            # Mostrar resumen por tabla
            log.info("\nResumen por tabla:")
            for _, row in df_tables.iterrows():
                log.info(f"  • {row['tabla']}: ~{row['filas_aprox']:,} filas, {row['total_kb']:.2f} KB")
    
    except Exception as e:
        log.exception(f"❌ Error generando diccionario: {e}")

# ---------- Ejecutar ----------
if __name__ == "__main__":
    start_time = time.time()
    
    try:
        # Ejecutar ETL
        run_etl()
        
        # Generar diccionario
        generar_diccionario()
        
        elapsed = time.time() - start_time
        log.info(f"\n{'='*60}")
        log.info(f"PROCESO COMPLETADO EN {elapsed:.2f} segundos")
        log.info(f"{'='*60}")
        
    except KeyboardInterrupt:
        log.warning("\n⚠ Proceso interrumpido por el usuario")
    except Exception as e:
        log.exception(f"❌ Error fatal: {e}")
    finally:
        engine.dispose()
        log.info("\n✓ Conexiones cerradas")